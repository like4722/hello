#!/usr/bin/python
# -*- coding: UTF-8 -*-
import pandas as pd
from openpyxl import load_workbook

#执管航班
dict_zhiguan = {'上海-成都': 'CZ3359', '成都-上海': 'CZ3360', '上海-昆明': 'CZ3677', '昆明-上海': 'CZ3678', '上海-珠海': 'CZ3680', '上海-广州': 'CZ3505', '广州-上海': 'CZ3506', '上海-芒市': 'CZ3913', '芒市-上海': 'CZ3914', '沈阳-上海': 'CZ6515', '上海-沈阳': 'CZ6516', '上海-汕头': 'CZ8263', '汕头-上海': 'CZ8264', '上海-西宁': 'CZ8839', '西宁-上海': 'CZ8840', '上海-深圳': 'CZ3562', '深圳-上海': 'CZ3589', '上海-台北': 'CZ3095', '台北-上海': 'CZ3096', '上海-大阪': 'CZ8105', '大阪-上海': 'CZ8106', '上海-首尔': 'CZ369', '首尔-上海': 'CZ370', '上海-东京': 'CZ6051', '东京-上海': 'CZ6052', '上海-胡志明': 'CZ6077', '胡志明-上海': 'CZ6078'}

#打开path路径的xlsx
list1 = []
list2 = []
def open_excel(path,sheet):
    df = pd.DataFrame(pd.read_excel(path, sheet))
    return df
    #print(df.values[0])
    """
    for i in range(df.iloc[:,0].size):
        list1.append(df.values[i][1])
        list2.append(df.values[i][0])
    dict_zhiguan = dict(zip(list1,list2))
    print(list1)
    print(list2)
    print(dict_zhiguan)
    """
def open_csv(path,sheet):
    df = pd.DataFrame(pd.read_csv(path, sheet))
    return df
    #print(df.values[0])
def write_excel(path,df):
    book = load_workbook(path)
    if 'Sheet1' in book.sheetnames:
        del book['Sheet1']
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    df.to_excel(writer, 'Sheet1', index=False)
    writer.save()
    writer.close()

def csv_to_xlsx_pd(path1,path2):
    csv = pd.read_csv(open(path1), encoding='utf-8')
    csv.to_excel(path2, sheet_name='Sheet1',index=False)

def main():
    path = r'C:\Users\like\Desktop\执管航班.xls'
    sheet ="航线"
    #open_excel(path,sheet)
    open_csv(path, sheet)
if __name__ == '__main__':
    main()
