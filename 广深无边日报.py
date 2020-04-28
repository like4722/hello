#-*- coding:utf-8 -*-
import numpy as np
import pandas as pd 
import openpyxl
import math
from openpyxl import load_workbook
from datetime import timedelta
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')
print('\n广深无边日报表\n')

def cal(x,hangxian):
    x_cal=pd.DataFrame()
    for i in range(len(hangxian)):
        ttt=x[x['航线中文']==hangxian[i]]
        if ttt.empty:
            continue
        else:
            x_cal.loc[i,'航线中文']=hangxian[i]
            x_cal.loc[i,'收入']=sum(ttt['收入(快报)'])/10000
            x_cal.loc[i,'座公里']=sum(ttt['座公里(航节)'])/10000
            x_cal.loc[i,'客座率']=sum(ttt['叠加登机数'])/sum(ttt['座位数(航节)'])
            x_cal.loc[i,'票价']=sum(ttt['收入(快报)'])/sum(ttt['登机数(快报)'])
            x_cal.loc[i,'座收']=sum(ttt['收入(快报捆绑)'])/sum(ttt['座公里(捆绑)'])
    return x_cal

def apanage(x):
    SHA=['舟山','井冈山','南昌','赣州','宁波','浦东','虹桥','上海',
    '温州','无锡','宜春','安庆','阜阳','合肥','九华山','池州','黄山']
    HGH=['杭州']
    NKG=['常州','南京','连云港','徐州','盐城','扬州']
    for i in range(len(x)):
        a=x.loc[i,'航线中文']
        a=a.split('-')
        a=a=[a[0],'']
        if set(a)&set(SHA):
            x.loc[i,'管辖单位中文']='中心本部'
        if set(a)&set(HGH):
            x.loc[i,'管辖单位中文']='杭州'
        if set(a)&set(NKG):
            x.loc[i,'管辖单位中文']='南京'
    xx=x['管辖单位中文']
    return xx
def marginal(x):
    a=x.sum(level='航线中文')
    a['边际贡献率']=a['边际贡献']/a['收入合计']
    a=pd.DataFrame(a,columns=['班次','边际贡献率'])
    return a

#a=input('请输入数据截止日期(八位数字：20190101）\n')
a=int(20190703)#数据截止日期

df=pd.read_csv('E:\\数据源\\广深数据源\\收入快报'
                +str(a//100)+'最新.csv',encoding='GBK',low_memory=False)
dfl=pd.read_csv('E:\\数据源\\广深数据源\\收入快报'
                +str(a//100-100)+'全.csv',encoding='GBK',low_memory=False)
#文件名修改处
dwb=pd.read_csv('E:\\数据源\\无边数据源\\无边'
                +str(a//100)+'最新.csv',encoding='GBK')
route=pd.read_csv('航线名.csv',encoding='GBK')

df=pd.DataFrame(df,columns=['飞行日期','合并承运人中文','航线中文',
    '收入(快报)','座公里(航节)','叠加登机数','座位数(航节)',
    '登机数(快报)','收入(快报捆绑)','座公里(捆绑)'])
dfl=pd.DataFrame(dfl,columns=['飞行日期','合并承运人中文','航线中文',
    '收入(快报)','座公里(航节)','叠加登机数','座位数(航节)',
    '登机数(快报)','收入(快报捆绑)','座公里(捆绑)'])
df['飞行日期']=pd.to_datetime(df['飞行日期'])
dfl['飞行日期']=pd.to_datetime(dfl['飞行日期'])
dwb['承运日期']=dwb['承运日期'].map(lambda x:datetime.strptime(str(x),'%Y%m%d'))
print('收入快报最新日期:',df.飞行日期.max().date())
print('边际贡献最新日期:',dwb.承运日期.max().date())

endday=datetime.strptime(str(a),'%Y%m%d')
monthstart=datetime.strptime(str(a//100*100+1),'%Y%m%d')
lastendday=datetime.strptime(str(a-10000),'%Y%m%d')
lastmonthstart=datetime.strptime(str(a//100*100+1-10000),'%Y%m%d')
ss=monthstart.strftime('%m.%d-')
se=endday.strftime('%m.%d')

cz19=df[(df['飞行日期']>=monthstart)&(df['飞行日期']<=endday)
    &(df['合并承运人中文']=='南航')]
cz19previousday=df[(df['飞行日期']>=monthstart)&(df['飞行日期']<=endday-timedelta(1))
    &(df['合并承运人中文']=='南航')]
cz18=dfl[(dfl['飞行日期']>=lastmonthstart)&(dfl['飞行日期']<=lastendday)
    &(dfl['合并承运人中文']=='南航')]
cz18previousday=dfl[(dfl['飞行日期']>=lastmonthstart)&(dfl['飞行日期']<=lastendday-timedelta(1))
    &(dfl['合并承运人中文']=='南航')]

routeGS=route.iloc[:,0]
cz19_cal=cal(cz19,routeGS);
cz18_cal=cal(cz18,routeGS)
cz19previousday_cal=cal(cz19previousday,routeGS)
cz18previousday_cal=cal(cz18previousday,routeGS)

final=pd.DataFrame(columns=['管辖单位中文','航线中文','客收产投','产投环比',
    '收入','收入同比','座公里','座公里同比','客座率',
    '客座率同比','客座率环比','票价','票价同比',
    '票价环比','座收','座收同比','座收环比'])

final['航线中文']=list(map(str,cz19_cal['航线中文']));
final['管辖单位中文']=apanage(final);
final['收入']=cz19_cal['收入'];final['座公里']=cz19_cal['座公里']
final['收入同比']=final['收入']/cz18_cal['收入']-1
final['座公里同比']=final['座公里']/cz18_cal['座公里']-1
final['客座率']=cz19_cal['客座率'];final['票价']=cz19_cal['票价']
final['客座率同比']=final['客座率']-cz18_cal['客座率']
final['客座率环比']=final['客座率']-cz19previousday_cal['客座率']
final['票价同比']=final['票价']/cz18_cal['票价']-1
final['票价环比']=final['票价']/cz19previousday_cal['票价']-1
final['座收']=cz19_cal['座收']
final['座收同比']=final['座收']/cz18_cal['座收']-1
final['座收环比']=final['座收']/cz19previousday_cal['座收']-1
final['客收产投']=final['收入同比']-final['座公里同比']
final['产投环比']=final['客收产投']-(cz19previousday_cal['收入']/
cz18previousday_cal['收入']-cz19previousday_cal['座公里']/cz18previousday_cal['座公里'])
final.index.name='*数据累计'+ss+se
final=final.sort_values(by='收入',ascending=False)
abc=final[(final.座收环比==0)&(final.票价环比==0)
&(final.客座率环比==0)|(final.票价环比.isnull())].index
for i in abc:
    name=final.loc[i,'航线中文']
    dd=cz19[cz19.航线中文==name]
    time=dd.飞行日期
    if len(time.unique())==1:
        final.loc[i,'aaa']='只有'+datetime.strftime(time.min(),'%m.%d')
    if len(time.unique())>1:
        final.loc[i,'aaa']='环比上一班期';
        dd=dd[dd.飞行日期<time.max()]
        final.loc[i,'客座率环比']=final.loc[i,'客座率']-sum(dd['叠加登机数'])/sum(dd['座位数(航节)'])
        final.loc[i,'票价环比']=final.loc[i,'票价']/(sum(dd['收入(快报)'])/sum(dd['登机数(快报)']))-1
        final.loc[i,'座收环比']=final.loc[i,'座收']/(sum(dd['收入(快报捆绑)'])/sum(dd['座公里(捆绑)']))-1


print('广深完成')

dwb=pd.DataFrame(dwb,columns=['承运日期','航线中文','班次','收入合计','边际贡献'])
dwb=dwb.set_index('航线中文')
wb=marginal(dwb[(dwb['承运日期']>=monthstart)&
        (dwb['承运日期']<=endday)])
wb_previous=marginal(dwb[(dwb['承运日期']>=monthstart)&
        (dwb['承运日期']<=endday-timedelta(1))])
wb['环比前一日']=wb['边际贡献率']-wb_previous['边际贡献率']
#固定添加查看项
a=['温州-郑州','南京-海口']
#需求查看项
b=['宁波-惠州']
wb1=wb.loc[a,:]
wb1=wb1.reset_index()

#标准查看项
wbfinal=wb[(wb['班次']>0)&(wb['边际贡献率']<0.05)]
wbfinal=wbfinal.sort_values(by='边际贡献率')
wbfinal=wbfinal.reset_index()
#额外和标准合并
wbfinal=pd.concat([wbfinal,wb1],ignore_index=True)
#处理管辖单位和补贴
wbfinal['管辖单位中文']=apanage(wbfinal)
wbfinal=wbfinal.dropna()
butie=route.iloc[:,2:4]
wbfinal=wbfinal.merge(butie,on='航线中文',how='left')
wbfinal.index.name='*边际贡献率低于5%（含5%）的航线，数据累计'+ss+se
wbfinal=pd.DataFrame(wbfinal,columns=['管辖单位中文','航线中文',
            '班次','边际贡献率','环比前一日','是否补贴航线'])
abc=wbfinal[wbfinal.环比前一日==0].index
for i in abc:
    name=wbfinal.loc[i,'航线中文']
    dd=dwb.loc[name,:]
    time=dd.承运日期
    if len(time.unique())==1:
        wbfinal.loc[i,'aaa']='只飞了一天'
    if len(time.unique())>1:
        wbfinal.loc[i,'aaa']='环比上一班期';
        dd=dd[dd.承运日期<time.max()]
        wbfinal.loc[i,'环比前一日']=wbfinal.loc[i,'边际贡献率']-sum(dd['边际贡献'])/sum(dd['收入合计'])

print('无边完成')

book=load_workbook(r'E:\报表\广深报表\(广深+无边)模板.xlsx')
if '广深数据源' in book.sheetnames:
	del book['广深数据源']
if '无边数据源' in book.sheetnames:
	del book['无边数据源']
writer=pd.ExcelWriter(r'E:\报表\广深报表\(广深+无边)模板.xlsx',engine='openpyxl')
writer.book=book
final.to_excel(writer,'广深数据源')
wbfinal.to_excel(writer,'无边数据源')
writer.save()
print('完成')