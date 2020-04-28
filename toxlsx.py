#!/usr/bin/python
# -*- coding: UTF-8 -*-

import numpy as np
import sys
import pandas as pd
import pymysql
import openpyxl
import math
import time
from openpyxl import load_workbook
import datetime
from mysql import connect_mysqlnew
import xlrd
import xlutils.copy
import xlsxwriter
import warnings
import math
import GUI as gui
warnings.filterwarnings('ignore')
print('\nFFDT航线诊断工具\n')


def write_csv(route,date1,date2,flag):
    database = "sales"
    data1 = datetime.datetime.strptime(date1, '%Y-%m-%d') #当前开始时间
    data2 = datetime.datetime.strptime(date2, '%Y-%m-%d') #当前结束时间
    data2 = datetime.timedelta(days=1) + data2
    data3 = data1 - datetime.timedelta(days=365) #前一年开始时间
    data4 = data2 - datetime.timedelta(days=365) #前一年结束时间
    if flag == 0:
        sql_new = "select * from income_dss_main where 航线中文='" + route + "'"+ " and 飞行日期>='" + data1.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"
        sql_old = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data3.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime("%Y-%m-%d %H:%M:%S") + "'"
        print(sql_new)
        print(sql_old)
        conn = pymysql.connect(host="10.243.26.240", user="like", password="229398", database=database, charset="utf8")
        df_new = pd.read_sql(sql_new, con=conn)
        df_old = pd.read_sql(sql_old, con=conn)
        conn.close()
        return df_new, df_old
    else :
        sql_new_time1 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data1.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"+" and 起飞时间>='"+ "00:00:00"+"'"+" and 起飞时间<'"+ "08:00:00"+ "'"
        sql_new_time2 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data1.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "08:00:00" + "'" + " and 起飞时间<'" + "14:00:00" + "'"
        sql_new_time3 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data1.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "14:00:00" + "'" + " and 起飞时间<'" + "18:00:00" + "'"
        sql_new_time4 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data1.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "18:00:00" + "'" + " and 起飞时间<'" + "21:00:00" + "'"
        sql_new_time5 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data1.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "21:00:00" + "'" + " and 起飞时间<'" + "24:00:00" + "'"
        sql_old_time1 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data3.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "00:00:00" + "'" + " and 起飞时间<'" + "08:00:00" + "'"
        sql_old_time2 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data3.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "08:00:00" + "'" + " and 起飞时间<'" + "14:00:00" + "'"
        sql_old_time3 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data3.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "14:00:00" + "'" + " and 起飞时间<'" + "18:00:00" + "'"
        sql_old_time4 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data3.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "18:00:00" + "'" + " and 起飞时间<'" + "21:00:00" + "'"
        sql_old_time5 = "select * from income_dss_main where 航线中文='" + route + "'" + " and 飞行日期>='" + data3.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime(
            "%Y-%m-%d %H:%M:%S") + "'" + " and 起飞时间>='" + "21:00:00" + "'" + " and 起飞时间<'" + "24:00:00" + "'"
        print(sql_new_time1)
        conn = pymysql.connect(host="10.243.26.240", user="like", password="229398", database=database, charset="utf8")
        df_new_time1= pd.read_sql(sql_new_time1, con=conn)
        df_new_time2 = pd.read_sql(sql_new_time2, con=conn)
        df_new_time3 = pd.read_sql(sql_new_time3, con=conn)
        df_new_time4 = pd.read_sql(sql_new_time4, con=conn)
        df_new_time5 = pd.read_sql(sql_new_time5, con=conn)
        df_old_time1 = pd.read_sql(sql_old_time1, con=conn)
        df_old_time2 = pd.read_sql(sql_old_time2, con=conn)
        df_old_time3 = pd.read_sql(sql_old_time3, con=conn)
        df_old_time4 = pd.read_sql(sql_old_time4, con=conn)
        df_old_time5 = pd.read_sql(sql_old_time5, con=conn)
        conn.close()
        return  df_new_time1,  df_new_time2,  df_new_time3,  df_new_time4,  df_new_time5,  df_old_time1,  df_old_time2,  df_old_time3,  df_old_time4,  df_old_time5
def write_csv_cabin(route,date1,date2,flag):
    database = "sales"
    data1 = datetime.datetime.strptime(date1, '%Y-%m-%d') #当前开始时间
    data2 = datetime.datetime.strptime(date2, '%Y-%m-%d') #当前结束时间
    data2 = datetime.timedelta(days=1) + data2

    data3 = data1 - datetime.timedelta(days=365) #前一年开始时间
    data4 = data2 - datetime.timedelta(days=365) #前一年结束时间
    if flag == 0:
        #sql_cabin_new = "select * from cabin_dss where 航线中文='" + route + "'"+ " and 飞行日期>='" + data1.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"
        #sql_cabin_old = "select * from cabin_dss where 航线中文='" + route + "'" + " and 飞行日期>='" + data3.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime("%Y-%m-%d %H:%M:%S") + "'"
        sql_cabin_new = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" +  " >='" + data1.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"+" and a.`航线中文`='" + route+ "'"
        sql_cabin_old = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" +  " >='" + data3.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime("%Y-%m-%d %H:%M:%S") + "'"+" and a.`航线中文`='" + route+ "'"
        print(sql_cabin_new)
        print(sql_cabin_old)

        conn = pymysql.connect(host="10.243.26.240", user="like", password="229398", database=database, charset="utf8")

        df_cabin_new = pd.read_sql(sql_cabin_new, con=conn)
        df_cabin_old = pd.read_sql(sql_cabin_old, con=conn)
        conn.close()
        return df_cabin_new,df_cabin_old
    else :
        sql_cabin_new_time1 ="select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" +  " >='" + data1.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"+" and a.`航线中文`='" + route+ "'" + " and 起飞时间>='" + "00:00:00" + "'" + " and 起飞时间<'" + "08:00:00" + "'"
        sql_cabin_new_time2 = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" +  " >='" + data1.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"+" and a.`航线中文`='" + route+ "'" + " and 起飞时间>='" + "08:00:00" + "'" + " and 起飞时间<'" + "14:00:00" + "'"
        sql_cabin_new_time3 ="select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" +  " >='" + data1.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"+" and a.`航线中文`='" + route+ "'" + " and 起飞时间>='" + "14:00:00" + "'" + " and 起飞时间<'" + "18:00:00" + "'"
        sql_cabin_new_time4 = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" +  " >='" + data1.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"+" and a.`航线中文`='" + route+ "'" + " and 起飞时间>='" + "18:00:00" + "'" + " and 起飞时间<'" + "21:00:00" + "'"
        sql_cabin_new_time5 ="select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" +  " >='" + data1.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data2.strftime("%Y-%m-%d %H:%M:%S") + "'"+" and a.`航线中文`='" + route+ "'" + " and 起飞时间>='" + "21:00:00" + "'" + " and 起飞时间<'" + "24:00:00" + "'"
        sql_cabin_old_time1 = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" + " >='" + data3.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and a.`航线中文`='" + route + "'" + " and 起飞时间>='" + "00:00:00" + "'" + " and 起飞时间<'" + "08:00:00" + "'"
        sql_cabin_old_time2 = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" + " >='" + data3.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and a.`航线中文`='" + route + "'" + " and 起飞时间>='" + "08:00:00" + "'" + " and 起飞时间<'" + "14:00:00" + "'"
        sql_cabin_old_time3 = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" + " >='" + data3.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and a.`航线中文`='" + route + "'" + " and 起飞时间>='" + "14:00:00" + "'" + " and 起飞时间<'" + "18:00:00" + "'"
        sql_cabin_old_time4 = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" + " >='" + data3.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and a.`航线中文`='" + route + "'" + " and 起飞时间>='" + "18:00:00" + "'" + " and 起飞时间<'" + "21:00:00" + "'"
        sql_cabin_old_time5 = "select * from cabin_dss1 as a right join cabin_corresponding AS b ON a.`子舱位` = b.`子舱位` and a.`合并承运人中文` = b.`承运人` where 飞行日期" + " >='" + data3.strftime("%Y-%m-%d %H:%M:%S") + "'" + " and 飞行日期<'" + data4.strftime( "%Y-%m-%d %H:%M:%S") + "'" + " and a.`航线中文`='" + route + "'" + " and 起飞时间>='" + "21:00:00" + "'" + " and 起飞时间<'" + "24:00:00" + "'"

        conn = pymysql.connect(host="10.243.26.240", user="like", password="229398", database=database, charset="utf8")
        df_cabin_new_time1 = pd.read_sql(sql_cabin_new_time1, con=conn)
        df_cabin_new_time2 = pd.read_sql(sql_cabin_new_time2, con=conn)
        df_cabin_new_time3 = pd.read_sql(sql_cabin_new_time3, con=conn)
        df_cabin_new_time4 = pd.read_sql(sql_cabin_new_time4, con=conn)
        df_cabin_new_time5 = pd.read_sql(sql_cabin_new_time5, con=conn)
        df_cabin_old_time1 = pd.read_sql(sql_cabin_old_time1, con=conn)
        df_cabin_old_time2 = pd.read_sql(sql_cabin_old_time2, con=conn)
        df_cabin_old_time3 = pd.read_sql(sql_cabin_old_time3, con=conn)
        df_cabin_old_time4 = pd.read_sql(sql_cabin_old_time4, con=conn)
        df_cabin_old_time5 = pd.read_sql(sql_cabin_old_time5, con=conn)
        conn.close()
        return  df_cabin_new_time1,df_cabin_new_time2,df_cabin_new_time3,df_cabin_new_time4,df_cabin_new_time5,df_cabin_old_time1,df_cabin_old_time2,df_cabin_old_time3,df_cabin_old_time4,df_cabin_old_time5

#数据提取
def data_comp(df,x,comp):
    #座位数
    cap = df[x][(df[x]['合并承运人中文'] == comp)]['座位数_航节'].sum()
    #登机数
    per = df[x][(df[x]['合并承运人中文'] == comp)]['登机数_快报'].sum()
    #座公里
    seatmiles = df[x][(df[x]['合并承运人中文'] == comp)]['座公里_航节'].sum()/10000
    #收入
    income = df[x][(df[x]['合并承运人中文'] == comp)]['收入_快报'].sum()/10000
    #叠加登机数
    addper = df[x][(df[x]['合并承运人中文'] == comp)]['叠加登机数'].sum()
    # 收入捆绑
    incomeshare = df[x][(df[x]['合并承运人中文'] == comp)]['收入_快报捆绑'].sum()
    # 座公里捆绑
    seatmilesshare = df[x][(df[x]['合并承运人中文'] == comp)]['座公里_捆绑'].sum()
    #客座率
    per1 = addper/cap
    #平均票价
    aver_price = income/per*10000
    #座收
    per_mile_income = incomeshare/seatmilesshare
    list_data = [cap, per,seatmiles,income,per1,aver_price,per_mile_income]
    return list_data

def data_comp_cabin(df,x,comp):
    #两舱
    cabin_liangcang = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='两舱'))]['登机数'].sum()
    #全价票
    cabin_quanjiapiao =df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='全价票'))]['登机数'].sum()
    # Y-8
    cabin_Y8 = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='8折(含)至Y不含'))]['登机数'].sum()
    # 8-6
    cabin_86 = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='6折(含)至8折'))]['登机数'].sum()
    # 6-5
    cabin_65 = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='5折(含)至6折'))]['登机数'].sum()
    # 5-4
    cabin_54 = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='4折(含)至5折'))]['登机数'].sum()
    # 特殊产品舱
    cabin_special = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='特殊产品舱'))]['登机数'].sum()
    # 国际中转舱
    cabin_inter = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='国际中转舱'))]['登机数'].sum()
    # 国内中转舱
    cabin_domes = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='国内中转舱'))]['登机数'].sum()
    # 团队舱
    cabin_group = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='团队舱'))]['登机数'].sum()
    # 免票
    cabin_free = df[x][(df[x]['合并承运人中文'] == comp)][((df[x]['舱位区间新']=='免票'))]['登机数'].sum()
    sum_cabin = cabin_liangcang+cabin_quanjiapiao+cabin_Y8+cabin_86+cabin_65 +cabin_54+cabin_special +cabin_inter+cabin_domes+cabin_group+cabin_free
    list_data = [cabin_liangcang/sum_cabin,cabin_quanjiapiao/sum_cabin,cabin_Y8/sum_cabin,cabin_86/sum_cabin,cabin_65/sum_cabin,cabin_54/sum_cabin,cabin_special /sum_cabin,cabin_inter/sum_cabin,cabin_domes/sum_cabin,cabin_group/sum_cabin,cabin_free/sum_cabin]
    return list_data

def data_cabin(df,x):
    #两舱
    cabin_liangcang = df[x][((df[x]['舱位区间新']=='两舱'))]['登机数'].sum()
    #全价票
    cabin_quanjiapiao = df[x][((df[x]['舱位区间新']=='全价票'))]['登机数'].sum()
    # Y-8
    cabin_Y8 = df[x][((df[x]['舱位区间新']=='8折(含)至Y不含'))]['登机数'].sum()
    # 8-6
    cabin_86 = df[x][((df[x]['舱位区间新']=='6折(含)至8折'))]['登机数'].sum()
    # 6-5
    cabin_65 = df[x][((df[x]['舱位区间新']=='5折(含)至6折'))]['登机数'].sum()
    # 5-4
    cabin_54 = df[x][((df[x]['舱位区间新']=='4折(含)至5折'))]['登机数'].sum()
    # 特殊产品舱
    cabin_special = df[x][((df[x]['舱位区间新']=='特殊产品舱'))]['登机数'].sum()
    # 国际中转舱
    cabin_inter =df[x][((df[x]['舱位区间新']=='国际中转舱'))]['登机数'].sum()
    # 国内中转舱
    cabin_domes = df[x][((df[x]['舱位区间新']=='国内中转舱'))]['登机数'].sum()
    # 团队舱
    cabin_group = df[x][((df[x]['舱位区间新']=='团队舱'))]['登机数'].sum()
    # 免票
    cabin_free = df[x][((df[x]['舱位区间新']=='免票'))]['登机数'].sum()
    sum_cabin = cabin_liangcang+cabin_quanjiapiao+cabin_Y8+cabin_86+cabin_65 +cabin_54+cabin_special +cabin_inter+cabin_domes+cabin_group+cabin_free
    list_data = [cabin_liangcang/sum_cabin,cabin_quanjiapiao/sum_cabin,cabin_Y8/sum_cabin,cabin_86/sum_cabin,cabin_65/sum_cabin,cabin_54/sum_cabin,cabin_special /sum_cabin,cabin_inter/sum_cabin,cabin_domes/sum_cabin,cabin_group/sum_cabin,cabin_free/sum_cabin]
    return list_data
def data(df,x):
    #座位数
    cap = df[x]['座位数_航节'].sum()
    #登机数
    per = df[x]['登机数_快报'].sum()
    #座公里
    seatmiles = df[x]['座公里_航节'].sum()/10000
    #收入
    income = df[x]['收入_快报'].sum()/10000
    #叠加登机数
    addper = df[x]['叠加登机数'].sum()
    # 收入捆绑
    incomeshare = df[x]['收入_快报捆绑'].sum()
    # 座公里捆绑
    seatmilesshare = df[x]['座公里_捆绑'].sum()
    #客座率
    per1 = addper/cap
    #平均票价
    aver_price = income/per*10000
    #座收
    per_mile_income = incomeshare/seatmilesshare
    list_data = [cap, per,seatmiles,income,per1,aver_price,per_mile_income]
    return list_data
def judge(p):
    if(p>0):
        return "增幅"
    elif(p<0):
        return "降幅"
    else:
        return "持平"
def judge_cabin(p):
    if(p>0):
        return "下降"
    elif(p<0):
        return "增长"
    else:
        return "持平"
def judge_abs(p):
    if (p > 0):
        return "高对标航司"+str("%.4f" % abs(p))
    elif (p < 0):
        return "低对标航司"+str("%.4f" % abs(p))
    else:
        return "与对标航司持平"
def temp_data_time(df_total_1,df_total_2,df_total_3,df_total_4,argv1,argv2,argv3):
    starttime = datetime.datetime.now()
    list_df00 = ['小结']
    df00 = pd.DataFrame(columns=list_df00)
    list_df0 =  ['时间段']
    df0 = pd.DataFrame(columns=list_df0)
    p = "时间段："+str(argv2)+"-"+str(argv3)
    df0 = df0.append(pd.DataFrame([p], columns=list_df0))
    # part1
    list_df1 = ['航线', '南航', '东上航', '国深航', '海航', '吉祥航', '其他', '定位', '竞争对手', '航线性质', '强弱势']
    df1 = pd.DataFrame(columns=list_df1)
    list_data1 = [] #南航整体
    for i in range(2):
        list_data1.append(data_comp(df_total_1, i ,"南航"))
    list_data2 = [] #东上航整体
    for i in range(2):
        list_data2.append(data_comp(df_total_1, i, "东上航"))
    list_data3 = []  # 国深航整体
    for i in range(2):
        list_data3.append(data_comp(df_total_1, i, "国深航"))
    list_data0 = [] #整体
    for i in range(2):
        list_data0.append(data(df_total_1, i ))
    df1_1 = pd.DataFrame([[argv1,list_data1[0][0]/list_data0[0][0],list_data2[0][0]/list_data0[0][0],list_data3[0][0]/list_data0[0][0],0,0,0,"核心航线","东上航","公务性质","弱势"]], columns=list_df1)
    df1 = df1.append(df1_1)
    p = str(argv1)+"航线为有公商务性质的弱势、核心航线,其主要竞争对手是东上航。"
    print(p)
    df00 = df00.append(pd.DataFrame([p], columns=list_df00))
    endtime = datetime.datetime.now()
    print((endtime - starttime).microseconds)
    print("完成PART1")
    global flag
    flag = 1
    # part2
    starttime = datetime.datetime.now()
    list_df2 = ['飞行年', '座位数', '登机数', '座公里', '收入', '客座率', '平均票价', '座收']
    df2 = pd.DataFrame(columns=list_df2)
    df2_1 = pd.DataFrame([["2019年",list_data1[0][0],list_data1[0][1],list_data1[0][2],list_data1[0][3],list_data1[0][4],list_data1[0][5],list_data1[0][6]]], columns=list_df2)
    df2_2 = pd.DataFrame([["同比", list_data1[0][0] / list_data1[1][0] - 1, list_data1[0][1] / list_data1[1][1] - 1,
                           list_data1[0][2] / list_data1[1][2] - 1, list_data1[0][3] / list_data1[1][3] - 1,
                           list_data1[0][4] - list_data1[1][4], list_data1[0][5] / list_data1[1][5] - 1,
                           list_data1[0][6] / list_data1[1][6] - 1]], columns=list_df2)
    df2_3 = pd.DataFrame([["2019年", list_data2[0][0], list_data2[0][1], list_data2[0][2], list_data2[0][3],
                           list_data2[0][4], list_data2[0][5], list_data2[0][6]]], columns=list_df2)
    df2_4 = pd.DataFrame([["同比", list_data2[0][0] / list_data2[1][0] - 1, list_data2[0][1] / list_data2[1][1] - 1,
                           list_data2[0][2] / list_data2[1][2] - 1, list_data2[0][3] / list_data2[1][3] - 1,
                           list_data2[0][4] - list_data2[1][4], list_data2[0][5] / list_data2[1][5] - 1,
                           list_data2[0][6] / list_data2[1][6] - 1]], columns=list_df2)
    df2_5 = pd.DataFrame([["2019年", list_data0[0][0], list_data0[0][1], list_data0[0][2], list_data0[0][3],
                           list_data0[0][4], list_data0[0][5], list_data0[0][6]]], columns=list_df2)
    df2_6 = pd.DataFrame([["同比", list_data0[0][0] / list_data0[1][0] - 1, list_data0[0][1] / list_data0[1][1] - 1,
                           list_data0[0][2] / list_data0[1][2] - 1, list_data0[0][3] / list_data0[1][3] - 1,
                           list_data0[0][4] - list_data0[1][4], list_data0[0][5] / list_data0[1][5] - 1,
                           list_data0[0][6] / list_data0[1][6] - 1]], columns=list_df2)

    df2 = df2.append(df2_1)
    df2 = df2.append(df2_2)
    df2 = df2.append(df2_3)
    df2 = df2.append(df2_4)
    df2 = df2.append(df2_5)
    df2 = df2.append(df2_6)

    #市场运力同比
    a = list_data0[0][0] / list_data0[1][0] - 1
    #市场客源同比
    b = list_data0[0][1] / list_data0[1][1] - 1
    #南航运力同比
    a1 = list_data1[0][0] / list_data1[1][0] - 1
    # 南航客座率同比
    c1 = list_data1[0][4] - list_data1[1][4]
    # 南航平均票价同比
    d1 = list_data1[0][5] / list_data1[1][5] - 1
    # 南航座收同比
    e1 = list_data1[0][6] / list_data1[1][6] - 1
    # 南航座收
    f1 = list_data1[0][6]
    # 东航座收
    g1 = list_data2[0][6]
    if(a>=0):
        if(b>=0):
            if(abs(a-b)<=0.01):
                p = "市场运力增幅"+str("%.1f" % abs(a*100))+"%,"+"客流增幅与运力增幅持平"+"。"
            elif(a-b>0.01):
                p = "市场运力增幅" +str("%.1f" % abs(a*100))+"%,"+ "其中运力增幅大于客流增幅"+"。"
            else:
                p = "市场运力增幅" +str("%.1f" % abs(a*100))+"%,"+"其中客流增幅大于运力增幅"+"。"
        else:
            if(a-b<=0.01):
                p = "市场运力增幅" +str("%.1f" % abs(a*100))+"%,"+"客流与运力同比基本持平"+"。"
            else:
                p = "市场运力略增" + str("%.1f" % abs(a*100))+"%,"+ "客流同比略降"+str("%.1f" % abs(b*100))+"%"+"。"
    else:
        if (b < 0):
            if (abs(a - b) <= 0.01):
                p = "市场运力降幅" +str("%.1f" % abs(a*100))+"%,"+ "客流降幅与运力降幅基本持平"+"。"
            elif (a - b > 0.01):
                p = "市场运力降幅"+str("%.1f" % abs(a*100))+"%,"+ "客流降幅大于运力降幅"+"。"
            else:
                p = "市场运力降幅" +str("%.1f" % abs(a*100))+"%,"+"运力降幅大于客流降幅"+"。"
        else:
            if (b - a <= 0.01):
                p = "市场运力略降a" +str("%.1f" % abs(a*100))+"%,"+ "客流同比略增"+str("%.1f" % abs(b*100))+"%"+"。"
            else:
                p = "市场运力降幅a" +str("%.1f" % abs(a*100))+"%,"+ "客流同比增幅"+str(abs(int(b*100)))+"%"+"。"

    if(a1>0):
        if(a1>a):
            p = p + "南航运力增幅" + str("%.1f" % abs(a1*100)) + "%," + "增幅大于市场"+"，"
        elif(a1==a):
            p = p + "南航运力增幅" +str("%.1f" % abs(a1*100)) + "%," + "增幅与市场持平" + "，"
        else:
            p = p + "南航运力增幅" + str("%.1f" % abs(a1*100)) + "%," + "增幅小于市场" + "，"
    elif(a1<0):
        if(a1>a):
            p = p + "南航运力降幅" +str("%.1f" % abs(a1*100)) + "%," + "降幅小于市场" + "，"
        elif(a1==a):
            p = p + "南航运力降幅" + str("%.1f" % abs(a1*100)) + "%," + "降幅与市场持平" + "，"
        else:
            p = p + "南航运力降幅" + str("%.1f" % abs(a1*100)) + "%," + "降幅大于市场" + "，"
    else:
        p = p + "南航运力同比持平" + "，"

    df00 = df00.append(pd.DataFrame([p], columns=list_df00))
    p = "客座率同比"+judge(c1)+ str("%.1f" % abs(c1*100)) + "%"+",平均票价同比"+judge(d1)+ str("%.1f" % abs(d1*100)) + "%"+"，座收"+str("%.4f" % abs(f1))+"同比"+judge(e1)+ str("%.1f" % abs(e1*100)) + "%"
    #p = p + ","+judge_abs(f1-g1)
    print(p)
    df00 = df00.append(pd.DataFrame([p], columns=list_df00))
    p = judge_abs(f1 - g1)
    print(p)
    df00 = df00.append(pd.DataFrame([p], columns=list_df00))
    endtime = datetime.datetime.now()
    print((endtime - starttime).microseconds)
    print("完成PART2")
    flag = 2

    # part3
    starttime = datetime.datetime.now()
    list_df3 = ['飞行年', '座位数', '登机数', '座公里', '收入', '客座率', '平均票价', '座收']
    df3 = pd.DataFrame(columns=list_df3)
    list_data4= []
    for i in range(10):
        list_data4.append(data_comp(df_total_2,i,"南航"))
    for i in range(10):
        list_data4.append(data_comp(df_total_2, i, "东上航"))
    for i in range(10):
        list_data4.append(data(df_total_2, i))
    for i in range(5):
        df3 = df3.append(pd.DataFrame([['2019年', list_data4[i][0], list_data4[i][1], list_data4[i][2], list_data4[i][3],
                           list_data4[i][4],list_data4[i][5],list_data4[i][6]]], columns=list_df3))
        df3 = df3.append(pd.DataFrame([['同比', list_data4[i][0]/list_data4[i+5][0]-1, list_data4[i][1]/list_data4[i+5][1]-1, list_data4[i][2]/list_data4[i+5][2]-1,list_data4[i][3]/list_data4[i+5][3]-1,
                           list_data4[i][4]-list_data4[i+5][4], list_data4[i][5]/list_data4[i+5][5]-1, list_data4[i][6]/list_data4[i+5][6]-1]], columns=list_df3))
        df3 = df3.append(pd.DataFrame([['2019年', list_data4[i+10][0], list_data4[i+10][1], list_data4[i+10][2], list_data4[i+10][3],
                                  list_data4[i+10][4], list_data4[i+10][5], list_data4[i+10][6]]], columns=list_df3))
        df3 = df3.append(pd.DataFrame([['同比', list_data4[i+10][0] / list_data4[i+10+5][0] - 1,list_data4[i+10][1] / list_data4[i +10+ 5][1] - 1,
                            list_data4[i+10][2] / list_data4[i +10+ 5][2] - 1,list_data4[i+10][3] / list_data4[i +10+ 5][3] - 1,list_data4[i+10][4] -list_data4[i +10+ 5][4], list_data4[i+10][5] / list_data4[i+10+5][5] - 1,list_data4[i+10][6] / list_data4[i +10+ 5][6] - 1]], columns=list_df3))

    list_adva = []
    list_adva1 = []
    list_adva2 = ["早早班","上午班","下午班","晚班","晚晚班","早早班","上午班","下午班","晚班","晚晚班"]
    list_adva3 = []
    list_adva4 = []
    list_adva5 = []
    list_adva6 = []

    list_adva.extend([list_data4[0][6],list_data4[1][6],list_data4[2][6],list_data4[3][6],list_data4[4][6],list_data4[10][6],list_data4[11][6],list_data4[12][6],list_data4[13][6],list_data4[14][6],list_data4[20][6],list_data4[21][6],list_data4[22][6],list_data4[23][6],list_data4[24][6]])
    for i in range(5):
        if(list_adva[i]<list_adva[i+10]):
            list_adva1.append("劣势航班")
        else:
            list_adva1.append("优势航班")
    for i in range(5):
        if (list_adva[i+5] < list_adva[i + 10]):
            list_adva1.append("劣势航班")
        else:
            list_adva1.append("优势航班")
    for i in  range(5):
        if(list_adva1[i]=="优势航班"):
            list_adva3.append(list_adva2[i])
        else:
            list_adva4.append(list_adva2[i])
    for i in  range(5):
        if(list_adva1[i+5]=="优势航班"):
            list_adva5.append(list_adva2[i+5])
        else:
            list_adva6.append(list_adva2[i+5])

    p = "南航的优势航班时段为"+str(list_adva3)+","+"南航的劣势航班时段为"+str(list_adva4)+","+"对标航司的优势航班时段为"+str(list_adva5)+","+"对标航司的劣势航班时段为"+str(list_adva6)
    print(p)
    df00 = df00.append(pd.DataFrame([p], columns=list_df00))
    endtime = datetime.datetime.now()
    print((endtime - starttime).microseconds)
    print("完成PART3")
    flag = 3

    # part4
    starttime = datetime.datetime.now()
    list_df4 = ['飞行年', '两舱', '全价票', '8折(含)至Y不含', '6折(含)至8折', '5折(含)至6折', '4折(含)至5折', '特殊产品舱', '国际中转舱', '国内中转舱', '团队舱', '免票']
    df4 = pd.DataFrame(columns=list_df4)
    list_data5 = []
    for i in range(2):
        list_data5.append(data_comp_cabin(df_total_3, i, "南航"))
    list_data6 = []
    for i in range(2):
        list_data6.append(data_comp_cabin(df_total_3, i, "东上航"))
    list_data7 = []
    for i in range(2):
        list_data7.append(data_cabin(df_total_3, i))
    df4 = df4.append(pd.DataFrame([['2019年', list_data5[0][0], list_data5[0][1], list_data5[0][2], list_data5[0][3],
                           list_data5[0][4],list_data5[0][5],list_data5[0][6],list_data5[0][7],list_data5[0][8],list_data5[0][9],list_data5[0][10]]], columns=list_df4))
    df4 = df4.append(pd.DataFrame([['同比', list_data5[0][0]-list_data5[1][0], list_data5[0][1]-list_data5[1][1], list_data5[0][2]-list_data5[1][2], list_data5[0][3]-list_data5[1][3],
                           list_data5[0][4]-list_data5[1][4],list_data5[0][5]-list_data5[1][5],list_data5[0][6]-list_data5[1][6],list_data5[0][7]-list_data5[1][7],list_data5[0][8]-list_data5[1][8],list_data5[0][9]-list_data5[1][9],list_data5[0][10]-list_data5[1][10]]], columns=list_df4))
    df4 = df4.append(pd.DataFrame([['2019年', list_data6[0][0], list_data6[0][1], list_data6[0][2], list_data6[0][3],
                                    list_data6[0][4], list_data6[0][5], list_data6[0][6], list_data6[0][7],
                                    list_data6[0][8], list_data6[0][9], list_data6[0][10]]], columns=list_df4))
    df4 = df4.append(pd.DataFrame([['同比', list_data6[0][0] - list_data6[1][0], list_data6[0][1] - list_data6[1][1],
                                    list_data6[0][2] - list_data6[1][2], list_data6[0][3] - list_data6[1][3],
                                    list_data6[0][4] - list_data6[1][4], list_data6[0][5] - list_data6[1][5],
                                    list_data6[0][6] - list_data6[1][6], list_data6[0][7] - list_data6[1][7],
                                    list_data6[0][8] - list_data6[1][8], list_data6[0][9] - list_data6[1][9],
                                    list_data6[0][10] - list_data6[1][10]]], columns=list_df4))
    df4 = df4.append(pd.DataFrame([['2019年', list_data7[0][0], list_data7[0][1], list_data7[0][2], list_data7[0][3],
                                    list_data7[0][4], list_data7[0][5], list_data7[0][6], list_data7[0][7],
                                    list_data7[0][8], list_data7[0][9], list_data7[0][10]]], columns=list_df4))
    df4 = df4.append(pd.DataFrame([['同比', list_data7[0][0] - list_data7[1][0], list_data7[0][1] - list_data7[1][1],
                                    list_data7[0][2] - list_data7[1][2], list_data7[0][3] - list_data7[1][3],
                                    list_data7[0][4] - list_data7[1][4], list_data7[0][5] - list_data7[1][5],
                                    list_data7[0][6] - list_data7[1][6], list_data7[0][7] - list_data7[1][7],
                                    list_data7[0][8] - list_data7[1][8], list_data7[0][9] - list_data7[1][9],
                                    list_data7[0][10] - list_data7[1][10]]], columns=list_df4))
    cz_hcabin = list_data5[0][0]-list_data5[1][0]+list_data5[0][1]-list_data5[1][1]+list_data5[0][2]-list_data5[1][2]
    cz_mcabin = list_data5[0][3]-list_data5[1][3]+list_data5[0][4]-list_data5[1][4]
    cz_lcabin = list_data5[0][5]-list_data5[1][5]+list_data5[0][6]-list_data5[1][6]+list_data5[0][7]-list_data5[1][7]+list_data5[0][8]-list_data5[1][8]+list_data5[0][9]-list_data5[1][9]+list_data5[0][10]-list_data5[1][10]

    comp_hcabin = list_data6[0][0] - list_data6[1][0] + list_data6[0][1] - list_data6[1][1] + list_data6[0][2] - \
                  list_data6[1][2]
    comp_mcabin = list_data6[0][3] - list_data6[1][3] + list_data6[0][4] - list_data6[1][4]
    comp_lcabin = list_data6[0][5] - list_data6[1][5] + list_data6[0][6] - list_data6[1][6] + list_data6[0][7] - \
                  list_data6[1][7] + list_data6[0][8] - list_data6[1][8] + list_data6[0][9] - list_data6[1][9] + \
                  list_data6[0][10] - list_data6[1][10]
    hcabin = list_data7[0][0] - list_data7[1][0] + list_data7[0][1] - list_data7[1][1] + list_data7[0][2] - \
                  list_data7[1][2]
    mcabin = list_data7[0][3] - list_data7[1][3] + list_data7[0][4] - list_data7[1][4]
    lcabin = list_data7[0][5] - list_data7[1][5] + list_data7[0][6] - list_data7[1][6] + list_data7[0][7] - \
                  list_data7[1][7] + list_data7[0][8] - list_data7[1][8] + list_data7[0][9] - list_data7[1][9] + \
                  list_data7[0][10] - list_data7[1][10]
    p = "南航高舱同比" + judge_cabin(cz_hcabin) +str("%.1f" % abs(cz_hcabin *100))+ "%"+ "," + "中高舱位同比" + judge_cabin(cz_mcabin) + str("%.1f" % abs(cz_mcabin *100))+ "%"+"," + "预售舱位同比" + judge_cabin(cz_lcabin)+str("%.1f" % abs(cz_lcabin *100))+ "%"+ ";"
    print(p)
    df00 = df00.append(pd.DataFrame([p], columns=list_df00))
    p =  "对标航司高舱同比" + judge_cabin(comp_hcabin) + str(
        "%.1f" % abs(comp_hcabin * 100)) + "%" + "," + "中高舱位同比" + judge_cabin(comp_mcabin) + str(
        "%.1f" % abs(comp_mcabin * 100)) + "%" + "," + "预售舱位同比" + judge_cabin(comp_lcabin) + "%" + str(
        "%.1f" % abs(comp_lcabin * 100)) + "%" + ";"
    print(p)
    df00 = df00.append(pd.DataFrame([p], columns=list_df00))
    p = "市场整体高舱同比" + judge_cabin(hcabin) + str(
        "%.1f" % abs(hcabin * 100)) + "%" + "," + "中高舱位同比" + judge_cabin(mcabin) + str(
        "%.1f" % abs(mcabin * 100)) + "%" + "," + "预售舱位同比" + judge_cabin(lcabin) + str("%.1f" % abs(lcabin * 100)) + "%"+ ";"
    print(p)
    df00 = df00.append(pd.DataFrame([p], columns=list_df00))

    endtime = datetime.datetime.now()
    print((endtime - starttime).microseconds)
    print("完成PART4")
    flag = 4

    # part5
    starttime = datetime.datetime.now()
    list_df5 = ['飞行年', '两舱', '全价票', '8折(含)至Y不含', '6折(含)至8折', '5折(含)至6折', '4折(含)至5折', '特殊产品舱', '国际中转舱', '国内中转舱', '团队舱', '免票']
    df5 = pd.DataFrame(columns=list_df5)
    list_data8 = []
    for i in range(10):
        list_data8.append(data_comp_cabin(df_total_4, i, "南航"))

    for i in range(10):
        list_data8.append(data_comp_cabin(df_total_4, i, "东上航"))
    for i in range(5):
        df5 = df5.append(pd.DataFrame([['2019年', list_data8[i][0], list_data8[i][1], list_data8[i][2], list_data8[i][3],
                                        list_data8[i][4], list_data8[i][5], list_data8[i][6], list_data8[i][7], list_data8[i][8], list_data8[i][9], list_data8[i][10]]], columns=list_df5))
        df5 = df5.append(pd.DataFrame([['同比', list_data8[i][0]- list_data8[i + 5][0],
                                        list_data8[i][1] - list_data8[i + 5][1],
                                        list_data8[i][2] - list_data8[i + 5][2],
                                        list_data8[i][3] - list_data8[i + 5][3],
                                        list_data8[i][4] - list_data8[i + 5][4],
                                        list_data8[i][5] - list_data8[i + 5][5],
                                        list_data8[i][6] - list_data8[i + 5][6],
                                        list_data8[i][7] - list_data8[i + 5][7],
                                        list_data8[i][8] - list_data8[i + 5][8],
                                        list_data8[i][9] - list_data8[i + 5][9],
                                        list_data8[i][10] - list_data8[i + 5][10]]], columns=list_df5))
        df5 = df5.append(pd.DataFrame([['2019年', list_data8[i + 10][0], list_data8[i + 10][1], list_data8[i + 10][2], list_data8[i + 10][3],
              list_data8[i + 10][4], list_data8[i + 10][5], list_data8[i + 10][6], list_data8[i + 10][7], list_data8[i + 10][8], list_data8[i + 10][9], list_data8[i + 10][10]]], columns=list_df5))
        df5 = df5.append(pd.DataFrame([['同比', list_data8[i+10][0] - list_data8[i + 15][0],
                                        list_data8[i+10][1] - list_data8[i + 15][1],
                                        list_data8[i+10][2] - list_data8[i + 15][2],
                                        list_data8[i+10][3] - list_data8[i + 15][3],
                                        list_data8[i+10][4] - list_data8[i + 15][4],
                                        list_data8[i+10][5] - list_data8[i + 15][5],
                                        list_data8[i+10][6] - list_data8[i + 15][6],
                                        list_data8[i+10][7] - list_data8[i + 15][7],
                                        list_data8[i+10][8] - list_data8[i + 15][8],
                                        list_data8[i+10][9] - list_data8[i + 15][9],
                                        list_data8[i+10][10] - list_data8[i + 15][10]]], columns=list_df5))
    endtime = datetime.datetime.now()
    print((endtime - starttime).microseconds)
    print("完成PART5")
    flag = 5
    # 写入模板
    book = load_workbook(r'D:\FFDT\FFDT2模板.xlsx')
    if 'Sheet2' in book.sheetnames:
        del book['Sheet2']
    writer = pd.ExcelWriter(r'D:\FFDT\FFDT2模板.xlsx', engine='openpyxl')
    writer.book = book

    df0.to_excel(writer, sheet_name='Sheet2', startrow=1, startcol=0, index=False)
    df1.to_excel(writer, sheet_name='Sheet2', startrow=3, startcol=0, index=False)
    df2.to_excel(writer, sheet_name='Sheet2', startrow=6, startcol=0, index=False)
    df3.to_excel(writer, sheet_name='Sheet2', startrow=14, startcol=0, index=False)
    df4.to_excel(writer, sheet_name='Sheet2', startrow=36, startcol=0, index=False)
    df5.to_excel(writer, sheet_name='Sheet2', startrow=46, startcol=0, index=False)
    df00.to_excel(writer, sheet_name='Sheet2', startrow=68, startcol=0, index=False)
    writer.save()
    print("写入完成")

def main(argv):
    df_total_1=write_csv(argv[0],argv[1],argv[2],0)
    df_total_2 = write_csv(argv[0],argv[1],argv[2], 1)
    df_total_3 = write_csv_cabin(argv[0], argv[1], argv[2], 0)
    df_total_4 = write_csv_cabin(argv[0], argv[1], argv[2], 1)
    temp_data_time(df_total_1,df_total_2,df_total_3,df_total_4,argv[0],argv[1],argv[2])


if __name__ == '__main__':
    """
    x = input("航线中文 如上海-广州：")
    y = input("起始日期 如2019-06-01：")
    z = input("截止日期 如2019-06-30：")
    argv = [x,y,z]
    print(argv)
    main(argv)
    """
    main(["上海-广州","2019-07-01","2019-07-10"])
