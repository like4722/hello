#!/usr/bin/python
# -*- coding: UTF-8 -*-
import pandas as pd
from openpyxl import load_workbook
from module import open_excel
import math
import time

#def open_excel(path):
#    df = pd.DataFrame(pd.read_excel(path, 'Sheet1'))
 #   print(df.values[3][3])


def data_solved():
    df = pd.DataFrame(pd.read_excel(r'D:\策略小组\数据\20200224日报\明折明扣\test.xlsx', 'Sheet1'))
    list_xiehang = []
    for i in range(df.iloc[:, 0].size):
        if(df.values[i][3]=='/' or df.values[i][3]=='/J/C/'):
            list_xiehang.append(i)
    df = df.drop(labels=list_xiehang, axis=0)
    df = df.drop_duplicates()
    df = df.reset_index(drop=True)

    df['航线'] =  None
    df['航司'] =  None
    df['舱位'] =  None
    df['公布运价'] =  None
    df['Y舱公布运价'] =  None
    df['折扣系数'] =  None
    listx = []
    list_column = []
    list_cabin1 = []
    list_hangxian = []

    for i in range(df.iloc[:,0].size):
        if('FD:' == str(df.values[i][0])[0:3]):
            df.values[i] = str(df.values[i][0])[3:9]
        df.values[i][3] = str(df.values[i][3])[0:4]
        df.values[i][3] = str(df.values[i][3]).replace('.','')
        df.values[i][2] = str(df.values[i][2])[0]
        print(df.values[i][2])
        df.values[i][11] = str(df.values[i][1])[0:2]
        df.values[i][12] = str(df.values[i][1]).split('/')[-1]
        if (df.values[i][2] == "/"):
            if(len(str(df.values[i-1][0]))==6):
                df.values[i][0] = df.values[i-1][0]
            else:
                df.values[i][0] = df.values[i - 6][0]
        df.values[i][10] = df.values[i][0]
        df.values[i][13] = df.values[i][3]
        #记录Y舱运价
        if(df.values[i][12] == 'Y'):
            listx.append([df.values[i][10],df.values[i][11],df.values[i][13]])
    for i in range(df.iloc[:,0].size):
        #Y舱
        for j in range(len(listx)):
            if((listx[j][0] == df.values[i][10])&(listx[j][1] == df.values[i][11])):
                df.values[i][14] = listx[j][2]
                print(df.values[i][13],df.values[i][14])
                df.values[i][15] = round(float(df.values[i][13])/float(df.values[i][14]),2)
    for i in range(df.iloc[:,0].size):
        if(df.values[i][2]!='/'):
            list_column.append(i)
    df = df.drop(labels = list_column,axis=0)
    df = df.reset_index(drop = True)

    #处理重复舱位
    for i in range(df.iloc[:,0].size):
        list_hangxian.append(df.values[i][10])
        if('1' in str(df.values[i][12])):
            list_cabin1.append(i-1)
    df = df.drop(labels = list_cabin1,axis=0)
    df = df.reset_index(drop = True)
    len_df = df.iloc[:,0].size

    #处理舱位1标识
    df['舱位'] = df['舱位'].str.replace('1','',regex=True)
    df['舱位'] = df['舱位'].str.replace('OW', '', regex=True)
    df['舱位'] = df['舱位'].str.replace('RT', '', regex=True)
    list_hangxiannew = []
    #航线
    list_hangxiannew = list({}.fromkeys(list_hangxian).keys())
    #特殊舱位
    df_special = pd.DataFrame(pd.read_excel(r'D:\策略小组\数据\20200224日报\明折明扣\test.xlsx', '特殊舱位'))
    list_special1 = []
    list_special2 = []
    list_special3 = []
    for i in range(df_special.iloc[:,0].size):
        list_special1.append(df_special.values[i][0])
        list_special2.append(df_special.values[i][1])
        list_special3.append(df_special.values[i][3])
    flag = 0
    for i in range(len(list_hangxiannew)):
        for j in range(54):
            df.loc[i+flag+len_df] = ['','','','','','','','','','',list_hangxiannew[i],list_special1[j],list_special2[j],'','',list_special3[j]]
            flag = flag+1


    book=load_workbook(r'D:\策略小组\数据\20200224日报\明折明扣\test.xlsx')
    if 'Sheet2' in book.sheetnames:
	    del book['Sheet2']
    writer=pd.ExcelWriter(r'D:\策略小组\数据\20200224日报\明折明扣\test.xlsx',engine='openpyxl')
    writer.book=book
    df.to_excel(writer,'Sheet2',index=False)
    writer.save()
    writer.close()

#识别非共飞航线
def non_commonroute():
    list_comp = []
    list_route = []
    list_route1 = []
    list_czroute = []
    list_nonczroute = []
    list_noncommonroute = []
    df = open_excel(r'D:\策略小组\数据\20200224日报\明折明扣\test.xlsx','result')
    #print(df.values[3][0])
    for i in range(df.iloc[:, 0].size):
        if df.values[i][1] not in list_comp:
            list_comp.append(df.values[i][1])
        list_route.append([str(df.values[i][0])[6:8],str(df.values[i][0])[0:6]])
    print(len(list_route))
    #list_comp.pop()
    #list_route.pop()
    for i in range(len(list_route)):
        if list_route[i] not in list_route1:
            list_route1.append(list_route[i])
    list_route = list_route1
    list_route.pop()
    print("所有航司+航线数量:",len(list_route))

    #CZ列表
    for i in range(len(list_route)):
        if ('CZ' in list_route[i][0]) and (list_route[i] not in list_czroute):
            list_czroute.append(list_route[i])
    #print(list_czroute)
    # CZ列表添加SHA PVG
    for i in range(len(list_czroute)):
        if ('SHA' in list_czroute[i][1]):
            temp = [list_czroute[i][0],list_czroute[i][1].replace('SHA','PVG')]
            if(temp not in list_czroute):
                list_czroute.append(temp)
        else:
            if 'PVG' in list_czroute[i][1]:
                temp = [list_czroute[i][0], list_czroute[i][1].replace('PVG', 'SHA')]
                if (temp not in list_czroute):
                    list_czroute.append(temp)
            else:
                if (list_czroute[i] not in list_czroute):
                    list_czroute.append(list_czroute[i])
    print("CZ航线数PVG+SHA",len(list_czroute))
    print(list_czroute)

    #判断是否为非共飞航线
    list_czroute1 = []
    for i in range(len(list_czroute)):
        list_czroute1.append(list_czroute[i][1])
    for i in range(len(list_route)):
        print(list_czroute[1])
        if list_route[i][0] in ('ZH','MU','CA','FM'):
                if (list_route[i][1] not in list_czroute1) and (list_route[i] not in list_nonczroute):
                    #print(list_route[i][1])
                    list_nonczroute.append(list_route[i])
    print(list_nonczroute)
    print(len(list_nonczroute))
    book = load_workbook(r'D:\策略小组\数据\20200224日报\明折明扣\test.xlsx')
    if 'Sheet3' in book.sheetnames:
        del book['Sheet3']
    writer = pd.ExcelWriter(r'D:\策略小组\数据\20200224日报\明折明扣\test.xlsx', engine='openpyxl')
    writer.book = book
    df_new = pd.DataFrame(list_nonczroute, columns=['航司', '航线'])
    df_new.to_excel(writer, 'Sheet3', index=False)
    writer.save()
    writer.close()


def main():
    #FD = data_solved()
    time1 = time.time()
    non_commonroute()
    time2 = time.time()
    print(time2-time1)



if __name__ == '__main__':
    main()




